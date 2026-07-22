from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QTableWidget, QTableWidgetItem, 
                             QFileDialog, QMessageBox, QHeaderView)
from PyQt6.QtCore import Qt
import json
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils import parse_s7_data, pack_s7_data
from ui.icons import get_custom_icon

class CompareWindow(QDialog):
    def __init__(self, parent=None, simulate=False, plc_client=None, parent_dbs_structures=None):
        super().__init__(parent)
        self.simulate = simulate
        self.plc_client = plc_client
        # Inherit DB structures from main window if available
        self.dbs_structures = parent_dbs_structures if parent_dbs_structures is not None else {}
        
        self.setWindowTitle("Report Comparativo Snapshot (Compare Snapshots)")
        self.setMinimumSize(800, 500)
        self.setWindowIcon(get_custom_icon("compare"))
        
        # Snapshot Data storage
        self.snap_a_data = {}  # {db_num: {"size": size, "data": bytearray}}
        self.snap_b_data = {}
        self.snap_a_name = "Snapshot A / Live PLC"
        self.snap_b_name = "Snapshot B / File"
        
        self.differences = []  # List of dicts representing changes
        
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # File selector header
        files_layout = QHBoxLayout()
        
        # Left Side (Source A)
        self.source_a_label = QLabel("<b>Origine A:</b> [Nessuno - Clicca per caricare]")
        self.source_a_label.setStyleSheet("border: 1px solid rgb(63, 63, 70); padding: 5px; border-radius: 4px;")
        files_layout.addWidget(self.source_a_label, 1)
        
        self.load_a_btn = QPushButton("Carica File A...")
        self.load_a_btn.clicked.connect(lambda: self.load_snapshot("A"))
        files_layout.addWidget(self.load_a_btn)
        
        self.live_a_btn = QPushButton("Usa PLC Live")
        self.live_a_btn.setIcon(get_custom_icon("monitor"))
        self.live_a_btn.clicked.connect(self.load_live_plc_a)
        if not self.plc_client or not self.plc_client.is_connected():
            self.live_a_btn.setEnabled(False)
        files_layout.addWidget(self.live_a_btn)
        
        layout.addLayout(files_layout)
        
        # File selector B header
        files_b_layout = QHBoxLayout()
        
        # Right Side (Source B)
        self.source_b_label = QLabel("<b>Origine B:</b> [Nessuno - Clicca per caricare]")
        self.source_b_label.setStyleSheet("border: 1px solid rgb(63, 63, 70); padding: 5px; border-radius: 4px;")
        files_b_layout.addWidget(self.source_b_label, 1)
        
        self.load_b_btn = QPushButton("Carica File B...")
        self.load_b_btn.clicked.connect(lambda: self.load_snapshot("B"))
        files_b_layout.addWidget(self.load_b_btn)
        
        self.live_b_btn = QPushButton("Usa PLC Live")
        self.live_b_btn.setIcon(get_custom_icon("monitor"))
        self.live_b_btn.clicked.connect(self.load_live_plc_b)
        if not self.plc_client or not self.plc_client.is_connected():
            self.live_b_btn.setEnabled(False)
        files_b_layout.addWidget(self.live_b_btn)
        
        layout.addLayout(files_b_layout)
        
        # Compare Button
        self.compare_btn = QPushButton("Compara e Genera Report")
        self.compare_btn.setIcon(get_custom_icon("compare"))
        self.compare_btn.setStyleSheet("font-weight: bold;")
        self.compare_btn.clicked.connect(self.run_comparison)
        self.compare_btn.setEnabled(False)
        layout.addWidget(self.compare_btn)
        
        # Table of Differences
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Data Block", "Offset / Var", "Tipo", "Valore Origine A", "Valore Origine B", "Dettaglio"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 100)
        self.table.setColumnWidth(1, 120)
        self.table.setColumnWidth(2, 80)
        self.table.setColumnWidth(3, 140)
        self.table.setColumnWidth(4, 140)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        layout.addWidget(self.table)
        
        # Bottom Export Actions
        export_layout = QHBoxLayout()
        
        self.summary_label = QLabel("Nessun confronto eseguito.")
        export_layout.addWidget(self.summary_label, 1)
        
        self.export_csv_btn = QPushButton("Esporta CSV...")
        self.export_csv_btn.clicked.connect(self.export_csv)
        self.export_csv_btn.setEnabled(False)
        export_layout.addWidget(self.export_csv_btn)
        
        self.export_xlsx_btn = QPushButton("Esporta Excel (XLSX)...")
        self.export_xlsx_btn.clicked.connect(self.export_excel)
        self.export_xlsx_btn.setEnabled(False)
        export_layout.addWidget(self.export_xlsx_btn)
        
        layout.addLayout(export_layout)

    def load_snapshot(self, target):
        filepath, _ = QFileDialog.getOpenFileName(
            self, f"Carica Snapshot {target}", "", "Snapshot Files (*.xlsx *.s7d *.json);;All Files (*)"
        )
        if not filepath:
            return
            
        try:
            snap_data = {}
            if filepath.endswith('.xlsx'):
                wb = openpyxl.load_workbook(filepath, data_only=True)
                for sheet_name in wb.sheetnames:
                    # Look for sheets that start with DB
                    if sheet_name.replace(" ", "").upper().startswith("DB"):
                        ws = wb[sheet_name]
                        db_val = ws.cell(row=2, column=1).value
                        if db_val is None:
                            continue
                        db_num = int(db_val)
                        size = int(ws.cell(row=2, column=2).value or 0)
                        
                        byte_data = bytearray(size)
                        
                        # Read variable rows starting from row 5
                        row_idx = 5
                        while True:
                            name_val = ws.cell(row=row_idx, column=1).value
                            if name_val is None:
                                break
                            dtype = str(ws.cell(row=row_idx, column=2).value or "").strip()
                            offset_val = ws.cell(row=row_idx, column=3).value
                            value = ws.cell(row=row_idx, column=4).value
                            
                            if offset_val is not None:
                                offset_str = str(offset_val).strip()
                                if "." in offset_str:
                                    parts = offset_str.split('.')
                                    byte_off = int(parts[0])
                                    bit_off = int(parts[1])
                                else:
                                    byte_off = int(float(offset_str)) if offset_str else 0
                                    bit_off = 0
                                    
                                pack_s7_data(dtype, value, byte_off, bit_off, byte_data)
                                
                            row_idx += 1
                            
                        snap_data[db_num] = {
                            "size": size,
                            "data": byte_data
                        }
            else:
                with open(filepath, 'r') as f:
                    data = json.load(f)
                    
                if "dbs" not in data:
                    raise ValueError("File non valido. Manca la chiave 'dbs' dello snapshot.")
                    
                for db_str, content in data["dbs"].items():
                    db_num = int(db_str)
                    hex_data = content.get("data", "")
                    byte_data = bytearray.fromhex(hex_data)
                    snap_data[db_num] = {
                        "size": content.get("size", len(byte_data)),
                        "data": byte_data
                    }
                
            file_name = filepath.split('/')[-1].split('\\')[-1]
            
            if target == "A":
                self.snap_a_data = snap_data
                self.snap_a_name = f"File: {file_name}"
                self.source_a_label.setText(f"<b>Origine A:</b> {self.snap_a_name}")
            else:
                self.snap_b_data = snap_data
                self.snap_b_name = f"File: {file_name}"
                self.source_b_label.setText(f"<b>Origine B:</b> {self.snap_b_name}")
                
            self.check_compare_ready()
            
        except Exception as e:
            QMessageBox.critical(self, "Errore Caricamento", f"Impossibile leggere il file snapshot:\n{str(e)}")

    def load_live_plc_a(self):
        if not self.plc_client or not self.plc_client.is_connected():
            QMessageBox.warning(self, "Non Connesso", "Non sei connesso ad alcun PLC live.")
            return
            
        try:
            self.source_a_label.setText("<b>Origine A:</b> [Lettura PLC live...]")
            self.source_a_label.repaint()
            
            # Fetch all DBs and their current values
            db_list = self.plc_client.list_dbs()
            snap_data = {}
            failed_dbs = []
            
            for db_num in db_list:
                size = self.plc_client.get_db_size(db_num)
                if size == 0:
                    continue
                try:
                    data = self.plc_client.read_db_bytes(db_num, size)
                    snap_data[db_num] = {
                        "size": size,
                        "data": data
                    }
                except Exception as db_err:
                    failed_dbs.append((db_num, str(db_err)))
                    
            if not snap_data:
                raise ValueError("Nessun Data Block è stato letto con successo dal PLC.")
                
            self.snap_a_data = snap_data
            self.snap_a_name = f"PLC Live ({self.plc_client.ip})"
            self.source_a_label.setText(f"<b>Origine A:</b> {self.snap_a_name}")
            
            if failed_dbs:
                error_msgs = "\n".join([f"- DB {db}: {err}" for db, err in failed_dbs])
                QMessageBox.warning(
                    self, "Lettura PLC Parziale",
                    f"Lettura completata. Caricate con successo {len(snap_data)} DB dal PLC.\n\n"
                    f"Tuttavia, i seguenti {len(failed_dbs)} DB non sono stati letti a causa di errori:\n{error_msgs}"
                )
                
            self.check_compare_ready()
            
        except Exception as e:
            self.source_a_label.setText("<b>Origine A:</b> [Errore PLC live]")
            QMessageBox.critical(self, "Errore Lettura PLC", f"Impossibile leggere i valori live del PLC:\n{str(e)}")

    def load_live_plc_b(self):
        if not self.plc_client or not self.plc_client.is_connected():
            QMessageBox.warning(self, "Non Connesso", "Non sei connesso ad alcun PLC live.")
            return
            
        try:
            self.source_b_label.setText("<b>Origine B:</b> [Lettura PLC live...]")
            self.source_b_label.repaint()
            
            # Fetch all DBs and their current values
            db_list = self.plc_client.list_dbs()
            snap_data = {}
            failed_dbs = []
            
            for db_num in db_list:
                size = self.plc_client.get_db_size(db_num)
                if size == 0:
                    continue
                try:
                    data = self.plc_client.read_db_bytes(db_num, size)
                    snap_data[db_num] = {
                        "size": size,
                        "data": data
                    }
                except Exception as db_err:
                    failed_dbs.append((db_num, str(db_err)))
                    
            if not snap_data:
                raise ValueError("Nessun Data Block è stato letto con successo dal PLC.")
                
            self.snap_b_data = snap_data
            self.snap_b_name = f"PLC Live ({self.plc_client.ip})"
            self.source_b_label.setText(f"<b>Origine B:</b> {self.snap_b_name}")
            
            if failed_dbs:
                error_msgs = "\n".join([f"- DB {db}: {err}" for db, err in failed_dbs])
                QMessageBox.warning(
                    self, "Lettura PLC Parziale",
                    f"Lettura completata. Caricate con successo {len(snap_data)} DB dal PLC.\n\n"
                    f"Tuttavia, i seguenti {len(failed_dbs)} DB non sono stati letti a causa di errori:\n{error_msgs}"
                )
                
            self.check_compare_ready()
            
        except Exception as e:
            self.source_b_label.setText("<b>Origine B:</b> [Errore PLC live]")
            QMessageBox.critical(self, "Errore Lettura PLC", f"Impossibile leggere i valori live del PLC:\n{str(e)}")

    def check_compare_ready(self):
        ready = len(self.snap_a_data) > 0 and len(self.snap_b_data) > 0
        self.compare_btn.setEnabled(ready)

    def run_comparison(self):
        self.differences = []
        self.table.setRowCount(0)
        
        # Get all unique DB numbers from both sources
        all_dbs = sorted(list(set(self.snap_a_data.keys()) | set(self.snap_b_data.keys())))
        
        for db_num in all_dbs:
            # Check if DB exists in both
            in_a = db_num in self.snap_a_data
            in_b = db_num in self.snap_b_data
            
            if not in_a:
                self.differences.append({
                    "db": db_num, "offset": "N/A", "type": "N/A",
                    "val_a": "Assente", "val_b": "Presente", "detail": "DB non presente nell'Origine A"
                })
                continue
            if not in_b:
                self.differences.append({
                    "db": db_num, "offset": "N/A", "type": "N/A",
                    "val_a": "Presente", "val_b": "Assente", "detail": "DB non presente nell'Origine B"
                })
                continue
                
            # DB exists in both, compare sizes
            db_a = self.snap_a_data[db_num]
            db_b = self.snap_b_data[db_num]
            
            if db_a["size"] != db_b["size"]:
                self.differences.append({
                    "db": db_num, "offset": "Size", "type": "Size",
                    "val_a": f"{db_a['size']} byte", "val_b": f"{db_b['size']} byte",
                    "detail": "Dimensione del Data Block differente"
                })
            
            # Perform value comparison
            data_a = db_a["data"]
            data_b = db_b["data"]
            
            # Check if we have a structure mapping for this DB. If not, generate default WORD mapping.
            structure = self.dbs_structures.get(db_num, [])
            if not structure:
                min_size = min(len(data_a), len(data_b))
                structure = []
                for offset in range(0, min_size - 1, 2):
                    structure.append({
                        "name": f"DBW{offset}",
                        "type": "WORD",
                        "offset": float(offset)
                    })
                if min_size % 2 != 0:
                    offset = min_size - 1
                    structure.append({
                        "name": f"DBB{offset}",
                        "type": "BYTE",
                        "offset": float(offset)
                    })
            
            # Perform structured comparison
            for var in structure:
                name = var.get("name", "Var")
                offset = var.get("offset", 0.0)
                dtype = var.get("type", "BYTE")
                
                # Split offset
                offset_str = str(offset)
                if "." in offset_str:
                    parts = offset_str.split('.')
                    byte_off = int(parts[0])
                    bit_off = int(parts[1])
                else:
                    byte_off = int(offset)
                    bit_off = 0
                    
                val_a = parse_s7_data(dtype, data_a, byte_off, bit_off)
                val_b = parse_s7_data(dtype, data_b, byte_off, bit_off)
                
                if val_a is None or val_b is None:
                    # Out of bounds or parsing error
                    continue
                    
                if val_a != val_b:
                    self.differences.append({
                        "db": db_num,
                        "offset": f"{offset} ({name})",
                        "type": dtype,
                        "val_a": str(val_a),
                        "val_b": str(val_b),
                        "detail": "Valore variabile modificato"
                    })
                        
        self.populate_table()

    def populate_table(self):
        diff_count = len(self.differences)
        self.summary_label.setText(f"Trovate {diff_count} differenze.")
        
        self.table.setRowCount(diff_count)
        for i, diff in enumerate(self.differences):
            self.table.setItem(i, 0, QTableWidgetItem(f"DB {diff['db']}"))
            self.table.setItem(i, 1, QTableWidgetItem(str(diff['offset'])))
            self.table.setItem(i, 2, QTableWidgetItem(str(diff['type'])))
            self.table.setItem(i, 3, QTableWidgetItem(str(diff['val_a'])))
            self.table.setItem(i, 4, QTableWidgetItem(str(diff['val_b'])))
            self.table.setItem(i, 5, QTableWidgetItem(str(diff['detail'])))
            
            # Color items that represent differences
            for col in range(6):
                item = self.table.item(i, col)
                if item:
                    item.setBackground(Qt.GlobalColor.yellow if "modificato" in diff['detail'].lower() else QColor("#FFCDD2"))
                    
        has_diffs = diff_count > 0
        self.export_csv_btn.setEnabled(has_diffs)
        self.export_xlsx_btn.setEnabled(has_diffs)

    def export_csv(self):
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Salva Report CSV", "Report_Differenze_S7.csv", "CSV Files (*.csv)"
        )
        if not filepath:
            return
            
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';')
                # Headers and Metadata
                writer.writerow(["REPORT COMPARATIVO VALORI PLC S7"])
                writer.writerow([f"Origine A (Riferimento):; {self.snap_a_name}"])
                writer.writerow([f"Origine B (Confronto):; {self.snap_b_name}"])
                writer.writerow([])
                writer.writerow(["DB", "Offset / Variabile", "Tipo", f"Valore in {self.snap_a_name}", f"Valore in {self.snap_b_name}", "Dettaglio"])
                
                for diff in self.differences:
                    writer.writerow([
                        f"DB {diff['db']}",
                        diff['offset'],
                        diff['type'],
                        diff['val_a'],
                        diff['val_b'],
                        diff['detail']
                    ])
                    
            QMessageBox.information(self, "Esportazione Completata", "Report CSV salvato con successo.")
        except Exception as e:
            QMessageBox.critical(self, "Errore di Scrittura", f"Impossibile scrivere il file CSV:\n{str(e)}")

    def export_excel(self):
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Salva Report Excel", "Report_Differenze_S7.xlsx", "Excel Files (*.xlsx)"
        )
        if not filepath:
            return
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Differenze S7"
            
            # Gridlines visible
            ws.views.sheetView[0].showGridLines = True
            
            # Styled fonts and colors
            title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
            header_font = Font(name="Segoe UI", size=10, bold=True, color="000000")
            regular_font = Font(name="Segoe UI", size=10)
            bold_font = Font(name="Segoe UI", size=10, bold=True)
            
            teal_fill = PatternFill(start_color="009999", end_color="009999", fill_type="solid") # Siemens Teal
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Title
            ws.merge_cells("A1:F1")
            ws["A1"] = "REPORT COMPARATIVO VALORI PLC S7"
            ws["A1"].font = title_font
            ws["A1"].fill = teal_fill
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 40
            
            # Metadata
            ws["A3"] = "Origine A (Riferimento):"
            ws["A3"].font = bold_font
            ws["B3"] = self.snap_a_name
            ws["B3"].font = regular_font
            
            ws["A4"] = "Origine B (Confronto):"
            ws["A4"].font = bold_font
            ws["B4"] = self.snap_b_name
            ws["B4"].font = regular_font
            
            # Headers
            headers = ["DB", "Offset / Variabile", "Tipo Dato", f"Valore in A", f"Valore in B", "Dettaglio Modifica"]
            ws.append([]) # Blank row 5
            ws.append(headers) # Row 6
            ws.row_dimensions[6].height = 25
            
            for col_idx in range(1, 7):
                cell = ws.cell(row=6, column=col_idx)
                cell.font = header_font
                cell.fill = gray_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                
            # Write rows
            for diff in self.differences:
                row_data = [
                    f"DB {diff['db']}",
                    diff['offset'],
                    diff['type'],
                    diff['val_a'],
                    diff['val_b'],
                    diff['detail']
                ]
                ws.append(row_data)
                row_idx = ws.max_row
                ws.row_dimensions[row_idx].height = 20
                
                # Apply styles
                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = regular_font
                    cell.border = thin_border
                    cell.fill = yellow_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1: continue # Skip merged title row
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            wb.save(filepath)
            QMessageBox.information(self, "Esportazione Completata", "Report Excel (.xlsx) salvato con successo.")
        except Exception as e:
            QMessageBox.critical(self, "Errore di Scrittura", f"Impossibile scrivere il file Excel:\n{str(e)}")
from PyQt6.QtGui import QColor
