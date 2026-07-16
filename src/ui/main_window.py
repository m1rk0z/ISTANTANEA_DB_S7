from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QLabel, QLineEdit, QSpinBox, QCheckBox, QPushButton, 
                             QSplitter, QTreeView, QTabWidget, QTableWidget, 
                             QTableWidgetItem, QFileDialog, QMessageBox, 
                             QHeaderView, QStatusBar, QComboBox, QDialog, QGridLayout)
from PyQt6.QtCore import Qt, QTimer, QThread, pyqtSignal
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QAction
import json
import datetime
import time

from plc_comm import PLCClient, PLCCommError
from ui.icons import get_custom_icon
from ui.nodes_dialog import NodesDialog
from ui.db_viewer import DBViewer
from ui.compare_window import CompareWindow

class ConnectWorker(QThread):
    connected = pyqtSignal(list, dict)
    failed = pyqtSignal(str)

    def __init__(self, plc_client, ip, rack, slot):
        super().__init__()
        self.plc_client = plc_client
        self.ip = ip
        self.rack = rack
        self.slot = slot

    def run(self):
        try:
            self.plc_client.connect(self.ip, self.rack, self.slot)
            dbs_list = self.plc_client.list_dbs()
            dbs_sizes = {}
            for db_num in dbs_list:
                dbs_sizes[db_num] = 0  # To be probed on-demand
            self.connected.emit(dbs_list, dbs_sizes)
        except Exception as e:
            self.failed.emit(str(e))

class DBScanWorker(QThread):
    progress = pyqtSignal(int, int)
    db_found = pyqtSignal(int, int)
    finished = pyqtSignal(list)

    def __init__(self, plc_client, start_val, end_val):
        super().__init__()
        self.plc_client = plc_client
        self.start_val = start_val
        self.end_val = end_val
        self.is_cancelled = False

    def run(self):
        found_dbs = []
        if self.plc_client.simulate:
            # Simulation mode scan
            total = self.end_val - self.start_val + 1
            mock_dbs = {1: 14, 2: 14, 3: 140, 4: 44, 5: 44, 6: 18, 7: 28, 8: 28, 9: 28, 10: 192, 11: 92, 12: 400, 13: 10, 40: 28, 500: 100, 1000: 50}
            for idx, db_num in enumerate(range(self.start_val, self.end_val + 1)):
                if self.is_cancelled:
                    break
                if idx % 100 == 0 or db_num == self.end_val:
                    self.progress.emit(db_num, self.end_val)
                if db_num in mock_dbs:
                    self.db_found.emit(db_num, mock_dbs[db_num])
                    found_dbs.append(db_num)
                time.sleep(0.0001)
            self.finished.emit(sorted(found_dbs))
            return

        import snap7
        import threading
        
        db_numbers = list(range(self.start_val, self.end_val + 1))
        total = len(db_numbers)
        
        # Spawn exactly 3 scanning threads to run in parallel (safe limit for S7-400 CP)
        num_threads = min(3, total)
        thread_segments = [[] for _ in range(num_threads)]
        for idx, db_num in enumerate(db_numbers):
            thread_segments[idx % num_threads].append(db_num)
            
        completed_lock = threading.Lock()
        completed_count = 0
        found_dbs_lock = threading.Lock()
        
        def thread_scan_target(assigned_dbs):
            nonlocal completed_count
            
            c = snap7.client.Client()
            try:
                c.connect(self.plc_client.ip, self.plc_client.rack, self.plc_client.slot)
            except Exception:
                # Update progress for connection failure
                with completed_lock:
                    completed_count += len(assigned_dbs)
                return

            for db_num in assigned_dbs:
                if self.is_cancelled:
                    break
                    
                try:
                    # Probe DB existence
                    c.db_read(db_num, 0, 1)
                    
                    # Success! Perform binary search to probe DB size
                    low = 1
                    high = 65535
                    detected = 0
                    while low <= high:
                        if self.is_cancelled:
                            break
                        mid = (low + high) // 2
                        try:
                            c.db_read(db_num, 0, mid)
                            detected = mid
                            low = mid + 1
                        except Exception:
                            high = mid - 1
                    
                    size = detected if detected > 0 else 100
                    with found_dbs_lock:
                        found_dbs.append(db_num)
                    self.db_found.emit(db_num, size)
                except Exception:
                    pass
                    
                with completed_lock:
                    completed_count += 1
                    if completed_count % 200 == 0 or completed_count == total:
                        self.progress.emit(completed_count, total)
                        
            try:
                c.disconnect()
            except Exception:
                pass

        # Spawn threads
        threads = []
        for segment in thread_segments:
            t = threading.Thread(target=thread_scan_target, args=(segment,))
            t.start()
            time.sleep(0.1)  # Stagger TCP connection handshakes to avoid CPU spikes on CP module
            threads.append(t)
            
        # Wait for all threads to complete
        for t in threads:
            t.join()
            
        self.finished.emit(sorted(found_dbs))

class ScanRangeDialog(QDialog):
    def __init__(self, parent=None, default_start=1, default_end=65535):
        super().__init__(parent)
        self.setWindowTitle("Imposta Intervallo Scansione")
        self.setMinimumWidth(380)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(12)
        
        # Info note
        info_label = QLabel(
            "<b>Nota sulla velocità:</b> La scansione di intervalli molto grandi (es. fino a 65535) "
            "può richiedere tempo a seconda della latenza di rete del PLC. "
            "Ti consigliamo di limitare l'intervallo a quello realmente utilizzato."
        )
        info_label.setWordWrap(True)
        info_label.setStyleSheet("color: #666666; font-size: 9pt;")
        layout.addWidget(info_label)
        
        # Form
        form_layout = QGridLayout()
        form_layout.setSpacing(8)
        
        form_layout.addWidget(QLabel("DB Iniziale:"), 0, 0)
        self.start_input = QSpinBox()
        self.start_input.setRange(1, 65535)
        self.start_input.setValue(default_start)
        self.start_input.setFixedHeight(28)
        form_layout.addWidget(self.start_input, 0, 1)
        
        form_layout.addWidget(QLabel("DB Finale:"), 1, 0)
        self.end_input = QSpinBox()
        self.end_input.setRange(1, 65535)
        self.end_input.setValue(default_end)
        self.end_input.setFixedHeight(28)
        form_layout.addWidget(self.end_input, 1, 1)
        
        layout.addLayout(form_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.ok_btn = QPushButton("Avvia Scansione")
        self.ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.ok_btn)
        
        self.cancel_btn = QPushButton("Annulla")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(btn_layout)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # Window properties
        self.setWindowTitle("IstanteS7 - Siemens S7 PLC Backup & Tool")
        self.setMinimumSize(1000, 650)
        self.setWindowIcon(get_custom_icon("plc"))
        
        # S7 Comm client
        self.plc_client = PLCClient(simulate=False)
        self.dbs_list = []      # List of DB numbers found on PLC
        self.dbs_sizes = {}     # {db_num: size}
        self.dbs_structures = {} # {db_num: [variables]}
        
        self.config_path = "config.json"
        self.profiles = []
        
        # Set up GUI layouts
        self.init_ui()
        
        # Load last connection and directory profiles
        self.load_config()
        
        # Heartbeat timer for connection status check
        self.hb_timer = QTimer(self)
        self.hb_timer.timeout.connect(self.check_connection_hb)
        self.hb_timer.start(1000)

    def init_ui(self):
        # Create Menu Bar
        self.create_menu_bar()
        
        # Create Tool Bar
        self.create_tool_bar()
        
        # Main Central Widget
        central = QWidget()
        central.setObjectName("centralWidget")
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        # Splitter Layout (Left Tree, Right Tabs)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(splitter, 1)
        
        # Left Tree Pane
        self.tree = QTreeView()
        self.tree_model = QStandardItemModel()
        self.tree_model.setHorizontalHeaderLabels(["Oggetti di Progetto"])
        self.tree.setModel(self.tree_model)
        self.tree.clicked.connect(self.on_tree_node_clicked)
        splitter.addWidget(self.tree)
        
        # Build initial Tree hierarchy
        self.init_project_tree()
        
        # Right Workspace Tabs Pane
        self.tabs = QTabWidget()
        splitter.addWidget(self.tabs)
        
        # Tab 1: DB Blocks list
        self.blocks_tab = QWidget()
        self.init_blocks_tab()
        self.tabs.addTab(self.blocks_tab, "Elenco Data Blocks")
        
        # Tab 2: Individual DB Viewer
        self.db_viewer_tab = DBViewer(
            plc_client=self.plc_client, 
            on_structures_changed=self.on_structures_changed
        )
        self.tabs.addTab(self.db_viewer_tab, "Editor Mappa DB (Monitor)")
        
        # Set split ratio (Tree 25%, Workspace 75%)
        splitter.setSizes([250, 750])
        
        # Status Bar
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.update_status_bar("Disconnesso.")

    def create_menu_bar(self):
        menubar = self.menuBar()
        
        # File Menu
        file_menu = menubar.addMenu("File")
        
        open_act = QAction("Apri File Snapshot...", self)
        open_act.setIcon(get_custom_icon("project"))
        open_act.triggered.connect(self.restore_from_file)
        file_menu.addAction(open_act)
        
        file_menu.addSeparator()
        
        exit_act = QAction("Esci", self)
        exit_act.triggered.connect(self.close)
        file_menu.addAction(exit_act)
        
        # PLC Menu
        plc_menu = menubar.addMenu("PLC")
        
        nodes_act = QAction("Nodi Accessibili...", self)
        nodes_act.setIcon(get_custom_icon("scan"))
        nodes_act.triggered.connect(self.open_accessible_nodes)
        plc_menu.addAction(nodes_act)
        
        plc_menu.addSeparator()
        
        self.connect_act = QAction("Connetti", self)
        self.connect_act.triggered.connect(self.toggle_connection)
        plc_menu.addAction(self.connect_act)
        
        compare_act = QAction("Compara Snapshots...", self)
        compare_act.setIcon(get_custom_icon("compare"))
        compare_act.triggered.connect(self.open_comparison_window)
        plc_menu.addAction(compare_act)
        
        plc_menu.addSeparator()
        
        import_symbols_act = QAction("Importa Tabella Simbolico (SEQ/SDF/CSV)...", self)
        import_symbols_act.triggered.connect(self.import_symbols_file)
        plc_menu.addAction(import_symbols_act)
        
        # Help Menu
        help_menu = menubar.addMenu("?")
        about_act = QAction("Informazioni su...", self)
        about_act.triggered.connect(self.show_about)
        help_menu.addAction(about_act)

    def create_tool_bar(self):
        toolbar = self.addToolBar("S7 Connection Toolbar")
        toolbar.setMovable(False)
        
        # Connection Controls
        toolbar.addWidget(QLabel("IP PLC: "))
        self.ip_input = QLineEdit("192.168.1.10")
        self.ip_input.setMinimumWidth(135)
        self.ip_input.setMaximumWidth(150)
        toolbar.addWidget(self.ip_input)
        
        toolbar.addWidget(QLabel(" Rack: "))
        self.rack_input = QComboBox()
        self.rack_input.addItems([str(i) for i in range(8)])
        self.rack_input.setCurrentText("0")
        self.rack_input.setMinimumWidth(75)
        self.rack_input.setMaximumWidth(90)
        self.rack_input.setFixedHeight(28)
        toolbar.addWidget(self.rack_input)
        
        toolbar.addWidget(QLabel(" Slot: "))
        self.slot_input = QComboBox()
        self.slot_input.addItems([str(i) for i in range(16)])
        self.slot_input.setCurrentText("2") # Default S7-300 is slot 2
        self.slot_input.setMinimumWidth(75)
        self.slot_input.setMaximumWidth(90)
        self.slot_input.setFixedHeight(28)
        toolbar.addWidget(self.slot_input)
        
        self.sim_check = QCheckBox("Simula")
        self.sim_check.setChecked(False)
        self.sim_check.stateChanged.connect(self.on_sim_changed)
        toolbar.addWidget(self.sim_check)
        
        self.connect_btn = QPushButton("Connetti")
        self.connect_btn.clicked.connect(self.toggle_connection)
        toolbar.addWidget(self.connect_btn)
        
        toolbar.addSeparator()
        
        # Address Book (Rubrica)
        toolbar.addWidget(QLabel("Rubrica: "))
        self.rubrica_combo = QComboBox()
        self.rubrica_combo.setMinimumWidth(150)
        self.rubrica_combo.currentIndexChanged.connect(self.on_profile_selected)
        toolbar.addWidget(self.rubrica_combo)
        
        self.save_profile_btn = QPushButton("Salva")
        self.save_profile_btn.clicked.connect(self.save_current_profile)
        toolbar.addWidget(self.save_profile_btn)
        
        self.manage_profile_btn = QPushButton("Gestisci")
        self.manage_profile_btn.clicked.connect(self.manage_profiles)
        toolbar.addWidget(self.manage_profile_btn)
        
        toolbar.addSeparator()
        
        # Action Shortcuts
        nodes_btn = QPushButton("Nodi Accessibili")
        nodes_btn.setIcon(get_custom_icon("scan"))
        nodes_btn.clicked.connect(self.open_accessible_nodes)
        toolbar.addWidget(nodes_btn)
        
        compare_btn = QPushButton("Confronta")
        compare_btn.setIcon(get_custom_icon("compare"))
        compare_btn.clicked.connect(self.open_comparison_window)
        toolbar.addWidget(compare_btn)

    def init_blocks_tab(self):
        layout = QVBoxLayout(self.blocks_tab)
        
        # Title and Select Actions
        action_layout = QHBoxLayout()
        action_layout.addWidget(QLabel("<b>Blocchi Dati (DB) rilevati nel PLC:</b>"))
        
        select_all_btn = QPushButton("Seleziona Tutti")
        select_all_btn.clicked.connect(lambda: self.set_table_selection(True))
        action_layout.addWidget(select_all_btn)
        
        select_none_btn = QPushButton("Deseleziona Tutti")
        select_none_btn.clicked.connect(lambda: self.set_table_selection(False))
        action_layout.addWidget(select_none_btn)
        
        # Manual Add Button
        self.add_manual_db_btn = QPushButton("Aggiungi DB Manuale...")
        self.add_manual_db_btn.clicked.connect(self.add_db_manually)
        action_layout.addWidget(self.add_manual_db_btn)
        
        # Scan Range Button
        self.scan_range_btn = QPushButton("Scansiona Intervallo DB...")
        self.scan_range_btn.clicked.connect(self.toggle_db_range_scan)
        action_layout.addWidget(self.scan_range_btn)
        
        # Import Symbols Button
        self.import_symbols_btn = QPushButton("Importa Simbolico...")
        self.import_symbols_btn.clicked.connect(self.import_symbols_file)
        action_layout.addWidget(self.import_symbols_btn)
        
        layout.addLayout(action_layout)
        
        # Table listing DBs
        self.blocks_table = QTableWidget()
        self.blocks_table.setColumnCount(4)
        self.blocks_table.setHorizontalHeaderLabels(["Backup", "Numero DB", "Dimensione (byte)", "Descrizione"])
        self.blocks_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.blocks_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.blocks_table.itemDoubleClicked.connect(self.on_block_row_double_clicked)
        layout.addWidget(self.blocks_table)
        
        # Backup actions
        backup_layout = QHBoxLayout()
        
        self.backup_selected_btn = QPushButton("Salva Snapshot (Selezionati)...")
        self.backup_selected_btn.setIcon(get_custom_icon("backup"))
        self.backup_selected_btn.setEnabled(False)
        self.backup_selected_btn.clicked.connect(self.backup_selected_dbs)
        backup_layout.addWidget(self.backup_selected_btn)
        
        self.backup_all_btn = QPushButton("Salva Snapshot (Tutti)...")
        self.backup_all_btn.setIcon(get_custom_icon("backup"))
        self.backup_all_btn.setEnabled(False)
        self.backup_all_btn.clicked.connect(self.backup_all_dbs)
        backup_layout.addWidget(self.backup_all_btn)
        
        self.restore_btn = QPushButton("Ripristina da File Snapshot...")
        self.restore_btn.setIcon(get_custom_icon("restore"))
        self.restore_btn.clicked.connect(self.restore_from_file)
        backup_layout.addWidget(self.restore_btn)
        
        layout.addLayout(backup_layout)

    def init_project_tree(self):
        self.tree_model.clear()
        self.tree_model.setHorizontalHeaderLabels(["Oggetti di Progetto"])
        
        # Root Node: Project
        self.root_node = QStandardItem("Progetto S7 (Backup & Restore)")
        self.root_node.setIcon(get_custom_icon("project"))
        self.root_node.setFlags(self.root_node.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.tree_model.appendRow(self.root_node)
        
        # PLC Node
        self.plc_node = QStandardItem("PLC Offline")
        self.plc_node.setIcon(get_custom_icon("plc"))
        self.plc_node.setFlags(self.plc_node.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.root_node.appendRow(self.plc_node)
        
        # Blocks Folder Node
        self.blocks_folder_node = QStandardItem("Blocchi")
        self.blocks_folder_node.setIcon(get_custom_icon("project"))
        self.blocks_folder_node.setFlags(self.blocks_folder_node.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.plc_node.appendRow(self.blocks_folder_node)
        
        self.tree.expandAll()

    def update_project_tree_online(self):
        # Update connection node name
        status_suffix = " (Simulazione)" if self.plc_client.simulate else ""
        self.plc_node.setText(f"PLC: {self.plc_client.ip}{status_suffix}")
        
        # Clear child nodes of Blocks folder
        self.blocks_folder_node.removeRows(0, self.blocks_folder_node.rowCount())
        
        # Add listed DBs as child nodes
        for db_num in self.dbs_list:
            db_item = QStandardItem(f"DB {db_num}")
            db_item.setIcon(get_custom_icon("db"))
            db_item.setData(db_num, Qt.ItemDataRole.UserRole)
            db_item.setFlags(db_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.blocks_folder_node.appendRow(db_item)
            
        self.tree.expandAll()

    def on_tree_node_clicked(self, index):
        item = self.tree_model.itemFromIndex(index)
        if not item:
            return
            
        text = item.text()
        db_num = item.data(Qt.ItemDataRole.UserRole)
        
        if text == "Blocchi" or item == self.blocks_folder_node:
            self.tabs.setCurrentIndex(0) # Go to DB blocks list
        elif db_num is not None:
            self.tabs.setCurrentIndex(1) # Go to individual DB mapper
            size = self.dbs_sizes.get(db_num, 0)
            if size <= 0:
                self.update_status_bar(f"Rilevamento dimensione DB {db_num} in corso...")
                try:
                    size = self.plc_client.get_db_size(db_num)
                except Exception:
                    size = 100
                self.dbs_sizes[db_num] = size
                for row in range(self.blocks_table.rowCount()):
                    row_db_item = self.blocks_table.item(row, 1)
                    if row_db_item and int(row_db_item.text()) == db_num:
                        self.blocks_table.item(row, 2).setText(str(size))
                        break
            self.db_viewer_tab.set_active_db(db_num, size)

    def on_block_row_double_clicked(self, item):
        row = item.row()
        db_num_item = self.blocks_table.item(row, 1)
        if db_num_item:
            db_num = int(db_num_item.text())
            size = self.dbs_sizes.get(db_num, 0)
            if size <= 0:
                self.update_status_bar(f"Rilevamento dimensione DB {db_num} in corso...")
                try:
                    size = self.plc_client.get_db_size(db_num)
                except Exception:
                    size = 100
                self.dbs_sizes[db_num] = size
                self.blocks_table.item(row, 2).setText(str(size))
            
            # Switch tabs and load viewer
            self.tabs.setCurrentIndex(1)
            self.db_viewer_tab.set_active_db(db_num, size)

    def on_sim_changed(self, state):
        simulate = (state == 2)
        if self.plc_client.is_connected():
            QMessageBox.warning(self, "Disconnetti prima", "Disconnetti la connessione corrente prima di cambiare modalità.")
            # Revert checkbox state
            self.sim_check.blockSignals(True)
            self.sim_check.setChecked(self.plc_client.simulate)
            self.sim_check.blockSignals(False)
            return
            
        self.plc_client = PLCClient(simulate=simulate)
        self.db_viewer_tab.plc_client = self.plc_client

    def toggle_connection(self):
        if self.plc_client.is_connected():
            self.plc_client.disconnect()
            self.on_disconnected()
        else:
            ip = self.ip_input.text().strip()
            rack = int(self.rack_input.currentText())
            slot = int(self.slot_input.currentText())
            
            # Clear old lists for the new connection
            self.blocks_table.setRowCount(0)
            self.dbs_list = []
            self.dbs_sizes = {}
            self.init_project_tree()
            
            self.update_status_bar(f"Connessione in corso a {ip}...")
            self.connect_btn.setEnabled(False)
            self.connect_btn.setText("Connessione...")
            
            # Disable configs during connection
            self.ip_input.setEnabled(False)
            self.rack_input.setEnabled(False)
            self.slot_input.setEnabled(False)
            self.sim_check.setEnabled(False)
            
            # Spawn background ConnectWorker thread
            self.connect_worker = ConnectWorker(self.plc_client, ip, rack, slot)
            self.connect_worker.connected.connect(self.on_connect_success)
            self.connect_worker.failed.connect(self.on_connect_failed)
            self.connect_worker.start()

    def on_connect_success(self, dbs_list, dbs_sizes):
        self.dbs_list = dbs_list
        self.dbs_sizes = dbs_sizes
        self.on_connected()
        
        # Automatically trigger background scan for all DBs (1-65535) if the PLC block listing returned nothing (e.g. S7-400 fallback)
        if not dbs_list and not self.plc_client.simulate:
            self.update_status_bar("Connesso. Avvio scansione automatica DB (1-65535) in corso...")
            self.scan_range_btn.setText("Ferma Scansione")
            self.db_scan_worker = DBScanWorker(self.plc_client, 1, 65535)
            self.db_scan_worker.progress.connect(self.on_db_scan_progress)
            self.db_scan_worker.db_found.connect(self.on_db_found)
            self.db_scan_worker.finished.connect(self.on_db_scan_finished)
            self.db_scan_worker.start()

    def on_connect_failed(self, error_msg):
        self.on_disconnected()
        QMessageBox.critical(self, "Errore di Connessione", f"Impossibile collegarsi al PLC:\n{error_msg}")

    def on_connected(self):
        self.connect_btn.setText("Disconnetti")
        self.connect_btn.setEnabled(True)
        self.connect_act.setText("Disconnetti")
        
        # Disable configs while connected
        self.ip_input.setEnabled(False)
        self.rack_input.setEnabled(False)
        self.slot_input.setEnabled(False)
        self.sim_check.setEnabled(False)
        
        # Enable actions
        self.backup_selected_btn.setEnabled(True)
        self.backup_all_btn.setEnabled(True)
        self.db_viewer_tab.monitor_checkbox.setEnabled(True)
        
        try:
            self.populate_blocks_table()
            self.update_project_tree_online()
            
            # Fetch CPU info if possible (non-blocking since already connected)
            cpu_info = {}
            try:
                cpu_info = self.plc_client.get_cpu_info()
            except Exception:
                pass
                
            if cpu_info:
                self.update_status_bar(
                    f"Connesso a {self.plc_client.ip} | CPU: {cpu_info.get('ModuleTypeName', 'N/A')} "
                    f"| Seriale: {cpu_info.get('SerialNumber', 'N/A')}"
                )
            else:
                self.update_status_bar(f"Connesso a {self.plc_client.ip} (Nessuna info CPU)")
        except Exception as e:
            QMessageBox.warning(self, "Errore Inizializzazione", f"Rilevato errore durante la configurazione della connessione:\n{str(e)}")
            self.update_status_bar(f"Connesso a {self.plc_client.ip}")

    def on_disconnected(self):
        if hasattr(self, 'db_scan_worker') and self.db_scan_worker.isRunning():
            self.db_scan_worker.is_cancelled = True
            self.db_scan_worker.wait()
            
        self.connect_btn.setText("Connetti")
        self.connect_btn.setEnabled(True)
        self.connect_act.setText("Connetti")
        
        # Re-enable inputs
        self.ip_input.setEnabled(True)
        self.rack_input.setEnabled(True)
        self.slot_input.setEnabled(True)
        self.sim_check.setEnabled(True)
        
        # Disable backup actions
        self.backup_selected_btn.setEnabled(False)
        self.backup_all_btn.setEnabled(False)
        
        # Stop DB viewer monitoring
        self.db_viewer_tab.monitor_checkbox.setChecked(False)
        self.db_viewer_tab.monitor_checkbox.setEnabled(False)
        
        self.update_status_bar("Disconnesso.")

    def check_connection_hb(self):
        # Periodically verifies if connection dropped in background
        if self.plc_client.is_connected():
            if not self.plc_client.simulate:
                # Active check
                try:
                    self.plc_client.client.get_connected()
                except Exception:
                    self.on_disconnected()
                    QMessageBox.warning(self, "Connessione Persa", "La connessione con il PLC è stata interrotta.")
        else:
            if self.connect_btn.text() == "Disconnetti":
                # State mismatch, handle it
                self.on_disconnected()

    def populate_blocks_table(self):
        # 1. Save checked DB numbers and identify already existing DBs
        checked_dbs = set()
        old_dbs = set()
        had_items = self.blocks_table.rowCount() > 0
        
        for row in range(self.blocks_table.rowCount()):
            db_item = self.blocks_table.item(row, 1)
            if db_item:
                db_num = int(db_item.text())
                old_dbs.add(db_num)
                chk_widget = self.blocks_table.cellWidget(row, 0)
                if chk_widget:
                    checkbox = chk_widget.layout().itemAt(0).widget()
                    if checkbox and checkbox.isChecked():
                        checked_dbs.add(db_num)
                        
        # Save current selected row's DB number
        selected_db = None
        selected_row_items = self.blocks_table.selectedItems()
        if selected_row_items:
            row = selected_row_items[0].row()
            db_num_item = self.blocks_table.item(row, 1)
            if db_num_item:
                selected_db = int(db_num_item.text())
                
        scroll_pos = self.blocks_table.verticalScrollBar().value()
        
        # 2. Re-populate table
        self.blocks_table.blockSignals(True)
        self.blocks_table.setRowCount(len(self.dbs_list))
        
        for i, db in enumerate(self.dbs_list):
            # Checkbox for backup selection
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            checkbox = QCheckBox()
            # New DBs are Checked (True) by default. Existing DBs preserve their Checked state.
            checkbox.setChecked(not had_items or db not in old_dbs or db in checked_dbs)
            chk_layout.addWidget(checkbox)
            self.blocks_table.setCellWidget(i, 0, chk_widget)
            
            # DB Number
            db_item = QTableWidgetItem(str(db))
            db_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            db_item.setIcon(get_custom_icon("db"))
            db_item.setFlags(db_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.blocks_table.setItem(i, 1, db_item)
            
            # Size
            size_item = QTableWidgetItem(str(self.dbs_sizes.get(db, 0)))
            size_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            size_item.setFlags(size_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.blocks_table.setItem(i, 2, size_item)
            
            # Description (editable)
            desc_val = f"Blocco Dati DB{db}"
            if hasattr(self, 'symbols_map') and db in self.symbols_map:
                sym_info = self.symbols_map[db]
                name = sym_info.get("name", "")
                comment = sym_info.get("comment", "")
                desc_val = f"{name} - {comment}" if comment else name
                
            desc_item = QTableWidgetItem(desc_val)
            self.blocks_table.setItem(i, 3, desc_item)
            
        # 3. Defer selection and scrollbar position restoration until layout updates are finished
        def restore_table_state():
            if selected_db is not None:
                for row in range(self.blocks_table.rowCount()):
                    db_item = self.blocks_table.item(row, 1)
                    if db_item and int(db_item.text()) == selected_db:
                        # Select items manually on each column item to avoid triggering QTableWidget auto-scroll
                        self.blocks_table.blockSignals(True)
                        for col in range(self.blocks_table.columnCount()):
                            cell_item = self.blocks_table.item(row, col)
                            if cell_item:
                                cell_item.setSelected(True)
                        self.blocks_table.blockSignals(False)
                        break
            self.blocks_table.verticalScrollBar().setValue(scroll_pos)
            
        QTimer.singleShot(0, restore_table_state)
        self.blocks_table.blockSignals(False)

    def set_table_selection(self, select):
        for row in range(self.blocks_table.rowCount()):
            chk_widget = self.blocks_table.cellWidget(row, 0)
            if chk_widget:
                checkbox = chk_widget.layout().itemAt(0).widget()
                if checkbox:
                    checkbox.setChecked(select)

    def get_selected_table_dbs(self):
        selected_dbs = []
        for row in range(self.blocks_table.rowCount()):
            chk_widget = self.blocks_table.cellWidget(row, 0)
            if chk_widget:
                checkbox = chk_widget.layout().itemAt(0).widget()
                if checkbox and checkbox.isChecked():
                    db_num_item = self.blocks_table.item(row, 1)
                    if db_num_item:
                        selected_dbs.append(int(db_num_item.text()))
        return selected_dbs

    def backup_selected_dbs(self):
        dbs_to_backup = self.get_selected_table_dbs()
        if not dbs_to_backup:
            QMessageBox.warning(self, "Nessun DB selezionato", "Seleziona almeno un DB da includere nel backup snapshot.")
            return
        self.execute_backup(dbs_to_backup)

    def backup_all_dbs(self):
        self.execute_backup(self.dbs_list)

    def execute_backup(self, dbs_list):
        if not self.plc_client.is_connected():
            QMessageBox.warning(self, "Non connesso", "Devi connetterti al PLC per effettuare il backup.")
            return
            
        filepath, selected_filter = QFileDialog.getSaveFileName(
            self, "Salva Snapshot S7", f"Snapshot_S7_{datetime.date.today().strftime('%Y%m%d')}.xlsx", 
            "Excel Snapshot Files (*.xlsx);;S7 Snapshot Files (*.s7d *.json)"
        )
        if not filepath:
            return
            
        self.update_status_bar("Esecuzione backup snapshot...")
        
        try:
            # Gather descriptions
            descriptions = {}
            for r in range(self.blocks_table.rowCount()):
                db_num_item = self.blocks_table.item(r, 1)
                if db_num_item:
                    db_num = int(db_num_item.text())
                    desc_item = self.blocks_table.item(r, 3)
                    descriptions[db_num] = desc_item.text() if desc_item else ""

            if filepath.endswith('.xlsx'):
                # Save to Excel Snapshot format
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "S7 Snapshot"
                
                # Write metadata
                ws["A1"] = "PLC IP"
                ws["B1"] = self.plc_client.ip
                ws["A2"] = "Timestamp"
                ws["B2"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws["A3"] = "Simulation"
                ws["B3"] = str(self.plc_client.simulate)
                
                # Write headers
                ws["A5"] = "DB"
                ws["B5"] = "Size"
                ws["C5"] = "DataHex"
                ws["D5"] = "Description"
                
                header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="008080", end_color="008080", fill_type="solid")
                for col in ["A", "B", "C", "D"]:
                    ws[f"{col}5"].font = header_font
                    ws[f"{col}5"].fill = header_fill
                    
                row_idx = 6
                for db in dbs_list:
                    size = self.dbs_sizes.get(db, 0)
                    if size == 0:
                        continue
                        
                    data = self.plc_client.read_db_bytes(db, size)
                    desc = descriptions.get(db, f"Blocco Dati DB{db}")
                    
                    ws.cell(row=row_idx, column=1, value=db)
                    ws.cell(row=row_idx, column=2, value=size)
                    ws.cell(row=row_idx, column=3, value=data.hex().upper())
                    ws.cell(row=row_idx, column=4, value=desc)
                    row_idx += 1
                    
                wb.save(filepath)
            else:
                # Save to legacy JSON/S7D format
                backup_data = {
                    "plc_ip": self.plc_client.ip,
                    "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "is_simulated": self.plc_client.simulate,
                    "dbs": {}
                }
                
                for db in dbs_list:
                    size = self.dbs_sizes.get(db, 0)
                    if size == 0:
                        continue
                        
                    data = self.plc_client.read_db_bytes(db, size)
                    desc = descriptions.get(db, f"Blocco Dati DB{db}")
                    
                    backup_data["dbs"][str(db)] = {
                        "size": size,
                        "data": data.hex().upper(),
                        "description": desc
                    }
                    
                with open(filepath, 'w') as f:
                    json.dump(backup_data, f, indent=4)
                    
            self.update_status_bar(f"Backup completato con successo: {len(dbs_list)} DB scritti.")
            QMessageBox.information(
                self, "Backup Completato",
                f"Backup di {len(dbs_list)} Data Blocks salvato correttamente in:\n{filepath}"
            )
        except Exception as e:
            self.update_status_bar("Backup fallito.")
            QMessageBox.critical(self, "Errore Backup", f"Si è verificato un errore durante la lettura/scrittura del backup:\n{str(e)}")

    def restore_from_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Apri Snapshot S7", "", "Snapshot Files (*.xlsx *.s7d *.json);;All Files (*)"
        )
        if not filepath:
            return
            
        try:
            backup_data = {}
            if filepath.endswith('.xlsx'):
                # Load from Excel
                import openpyxl
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
                
                backup_data = {
                    "plc_ip": ws["B1"].value,
                    "timestamp": ws["B2"].value,
                    "is_simulated": str(ws["B3"].value).lower() == "true",
                    "dbs": {}
                }
                
                row_idx = 6
                while True:
                    db_val = ws.cell(row=row_idx, column=1).value
                    if db_val is None:
                        break
                    db_num = int(db_val)
                    size = int(ws.cell(row=row_idx, column=2).value or 0)
                    data_hex = str(ws.cell(row=row_idx, column=3).value or "").strip()
                    desc = str(ws.cell(row=row_idx, column=4).value or "")
                    
                    backup_data["dbs"][str(db_num)] = {
                        "size": size,
                        "data": data_hex,
                        "description": desc
                    }
                    row_idx += 1
            else:
                # Load from legacy JSON/S7D
                with open(filepath, 'r') as f:
                    backup_data = json.load(f)
                    
            if "dbs" not in backup_data:
                raise ValueError("Il file snapshot non è nel formato corretto.")
                
            dbs_in_file = list(backup_data["dbs"].keys())
            
            # Show a dialog with checklist of DBs to restore
            from PyQt6.QtWidgets import QDialog, QListWidget, QListWidgetItem
            
            dialog = QDialog(self)
            dialog.setWindowTitle("Ripristina Snapshot")
            dialog.setMinimumSize(400, 300)
            
            diag_layout = QVBoxLayout(dialog)
            diag_layout.addWidget(QLabel("<b>Seleziona i Data Block da ripristinare sul PLC:</b>"))
            
            list_widget = QListWidget()
            for db_str in dbs_in_file:
                db_num = int(db_str)
                size = backup_data["dbs"][db_str].get("size", 0)
                item = QListWidgetItem(f"DB {db_num} ({size} byte)")
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(Qt.CheckState.Checked)
                item.setData(Qt.ItemDataRole.UserRole, db_num)
                list_widget.addItem(item)
                
            diag_layout.addWidget(list_widget)
            
            btn_layout = QHBoxLayout()
            ok_btn = QPushButton("Ripristina Selezionati nel PLC")
            ok_btn.clicked.connect(dialog.accept)
            btn_layout.addWidget(ok_btn)
            
            cancel_btn = QPushButton("Annulla")
            cancel_btn.clicked.connect(dialog.reject)
            btn_layout.addWidget(cancel_btn)
            
            diag_layout.addLayout(btn_layout)
            
            if dialog.exec() != QDialog.DialogCode.Accepted:
                return
                
            # Collect selected DBs to restore
            selected_dbs_restore = []
            for row in range(list_widget.count()):
                item = list_widget.item(row)
                if item.checkState() == Qt.CheckState.Checked:
                    selected_dbs_restore.append(item.data(Qt.ItemDataRole.UserRole))
                    
            if not selected_dbs_restore:
                QMessageBox.warning(self, "Nessun blocco", "Nessun DB selezionato per il ripristino.")
                return
                
            # Perform connection check
            if not self.plc_client.is_connected():
                QMessageBox.warning(self, "Non Connesso", "Devi connetterti al PLC prima di poter scrivere i dati.")
                return
                
            confirm = QMessageBox.question(
                self, "Conferma Ripristino",
                f"ATTENZIONE: Si sta per sovrascrivere la memoria di {len(selected_dbs_restore)} DB sul PLC live!\n"
                "Questa operazione modificherà i valori in esecuzione. Vuoi procedere?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if confirm == QMessageBox.StandardButton.No:
                return
                
            # Write bytes to PLC
            self.update_status_bar("Scrittura ripristino nel PLC in corso...")
            success_count = 0
            
            # Apply descriptions from the snapshot file back into our symbols_map
            if not hasattr(self, 'symbols_map'):
                self.symbols_map = {}
                
            for db in selected_dbs_restore:
                db_str = str(db)
                db_info = backup_data["dbs"][db_str]
                hex_data = db_info["data"]
                byte_data = bytearray.fromhex(hex_data)
                
                self.plc_client.write_db_bytes(db, byte_data)
                success_count += 1
                
                # Restore description if present in file
                if "description" in db_info and db_info["description"]:
                    self.symbols_map[db] = {
                        "name": db_info["description"],
                        "comment": ""
                    }
                    
            # Update blocks table to reflect loaded descriptions
            self.populate_blocks_table()
            
            self.update_status_bar(f"Ripristino completato: {success_count} DB riscritti nel PLC.")
            QMessageBox.information(
                self, "Ripristino Completato",
                f"Ripristino di {success_count} DB eseguito con successo sul PLC."
            )
            
            # Refresh live view if viewing the same DB
            if self.tabs.currentIndex() == 1 and self.db_viewer_tab.db_number in selected_dbs_restore:
                self.db_viewer_tab.read_full_db()
                
        except Exception as e:
            self.update_status_bar("Ripristino fallito.")
            QMessageBox.critical(self, "Errore Ripristino", f"Impossibile completare il ripristino:\n{str(e)}")

    def import_symbols_file(self):
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Importa Simbolico PLC (STEP 7)", "", "STEP 7 Symbol Files (*.asc);;All Symbol Files (*.asc *.seq *.sdf *.csv *.txt);;All Files (*)"
        )
        if not filepath:
            return
            
        try:
            symbols_map = {} # {db_num: {"name": symbol_name, "comment": comment}}
            
            # Read file with encoding fallback (ANSI/latin-1 is common for STEP 7 exports)
            lines = []
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
            except UnicodeDecodeError:
                with open(filepath, 'r', encoding='latin-1') as f:
                    lines = f.readlines()
                    
            import re
            for line in lines:
                # Skip comments or empty lines
                line = line.strip()
                if not line or line.startswith('//'):
                    continue
                    
                # Strip prefix number followed by comma (e.g., "126,")
                if ',' in line:
                    parts = line.split(',', 1)
                    content = parts[1].strip()
                else:
                    content = line
                    
                # Check for comments at the end of the line (often in quotes)
                comment = ""
                if content.count('"') >= 2 and content.endswith('"'):
                    last_quote = content.rfind('"')
                    first_quote = content.rfind('"', 0, last_quote)
                    if first_quote != -1:
                        comment = content[first_quote+1:last_quote]
                        content = content[:first_quote].strip()
                        
                # Split columns by tab or by 2 or more spaces
                if '\t' in content:
                    cols = [c.strip().strip('"') for c in content.split('\t') if c.strip()]
                else:
                    cols = [c.strip().strip('"') for c in re.split(r'\s{2,}', content) if c.strip()]
                    
                if len(cols) < 2:
                    continue
                    
                symbol_name = cols[0]
                address_type = ""
                address_num_str = ""
                
                if len(cols) >= 3:
                    if cols[1].upper() == "DB":
                        address_type = "DB"
                        address_num_str = cols[2]
                        
                # Handle combined address representation like "DB 3401"
                if not address_type and len(cols) >= 2:
                    addr_part = cols[1].replace(" ", "").upper()
                    if addr_part.startswith("DB") and addr_part[2:].isdigit():
                        address_type = "DB"
                        address_num_str = addr_part[2:]
                        
                if address_type == "DB" and address_num_str.isdigit():
                    db_num = int(address_num_str)
                    symbols_map[db_num] = {
                        "name": symbol_name,
                        "comment": comment
                    }
            
            if not symbols_map:
                QMessageBox.warning(self, "Nessun Simbolo Trovato", "Non è stato trovato alcun simbolo DB compatibile nel file selezionato.")
                return
                
            # Apply symbols to blocks list descriptions
            if not hasattr(self, 'symbols_map'):
                self.symbols_map = {}
            self.symbols_map.update(symbols_map)
            
            # Re-populate blocks table to show the new descriptions!
            self.populate_blocks_table()
            
            QMessageBox.information(
                self, "Importazione Completata",
                f"Importazione completata con successo! Trovati e applicati {len(symbols_map)} simboli DB."
            )
        except Exception as e:
            QMessageBox.critical(self, "Errore Importazione", f"Errore durante l'importazione del simbolico:\n{str(e)}")

    def open_accessible_nodes(self):
        self.nodes_dialog = NodesDialog(self, simulate=self.plc_client.simulate)
        if self.nodes_dialog.exec() == QDialog.DialogCode.Accepted:
            # Load found IP into toolbar
            self.ip_input.setText(self.nodes_dialog.selected_ip)
            self.rack_input.setCurrentText(str(self.nodes_dialog.selected_rack))
            self.slot_input.setCurrentText(str(self.nodes_dialog.selected_slot))
            
            # Defer auto-connect to prevent Event Loop conflicts and PyQt C++ crashes
            QTimer.singleShot(100, self.auto_connect_after_dialog)

    def auto_connect_after_dialog(self):
        if not self.plc_client.is_connected():
            self.toggle_connection()

    def open_comparison_window(self):
        # We pass self.dbs_structures so CompareWindow can show structured variables if defined
        self.compare_dialog = CompareWindow(
            self, 
            simulate=self.plc_client.simulate, 
            plc_client=self.plc_client,
            parent_dbs_structures=self.dbs_structures
        )
        self.compare_dialog.exec()

    def on_structures_changed(self, db_num, variables):
        # Store variable mappings globally so they persist and are shared with the CompareWindow
        self.dbs_structures[db_num] = variables

    def update_status_bar(self, text):
        sim_indicator = "[SIMULAZIONE] " if self.plc_client.simulate else ""
        self.status.showMessage(f"Stato: {sim_indicator}{text}")

    def show_about(self):
        QMessageBox.about(
            self, "Informazioni su IstanteS7",
            "<h3>IstanteS7 v1.0.0</h3>"
            "<p>Applicazione portable per il backup, ripristino e monitoraggio live "
            "dei Data Blocks (DB) per PLC Siemens Simatic S7-300 e S7-400.</p>"
            "<p>Stile grafico ispirato a <i>Siemens Simatic Manager Step 7</i>.</p>"
            "<p>Sviluppato in Python con PyQt6 e python-snap7 pure-python library.</p>"
        )

    def closeEvent(self, event):
        # Save last connection configuration on window close
        self.save_config()
        super().closeEvent(event)

    def load_config(self):
        import os
        self.profiles = []
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, 'r') as f:
                    config = json.load(f)
                    
                # Last connection details
                last_conn = config.get("last_connection", {})
                if last_conn:
                    self.ip_input.setText(last_conn.get("ip", "192.168.1.10"))
                    self.rack_input.setCurrentText(str(last_conn.get("rack", 0)))
                    self.slot_input.setCurrentText(str(last_conn.get("slot", 2)))
                    self.sim_check.setChecked(last_conn.get("simulate", False))
                    
                # Directory profiles
                self.profiles = config.get("profiles", [])
                if not self.profiles:
                    self.profiles = [
                        {"name": "Simulatore Locale", "ip": "127.0.0.1", "rack": 0, "slot": 2, "simulate": True},
                        {"name": "CPU Test (200.100.0.11)", "ip": "200.100.0.11", "rack": 0, "slot": 3, "simulate": False}
                    ]
                self.update_rubrica_combo()
            except Exception as e:
                print(f"Error loading config: {e}")
        else:
            # Default profiles
            self.profiles = [
                {"name": "Simulatore Locale", "ip": "127.0.0.1", "rack": 0, "slot": 2, "simulate": True},
                {"name": "CPU Test (200.100.0.11)", "ip": "200.100.0.11", "rack": 0, "slot": 3, "simulate": False}
            ]
            self.update_rubrica_combo()
            self.save_config()

    def save_config(self):
        try:
            config = {
                "last_connection": {
                    "ip": self.ip_input.text().strip(),
                    "rack": int(self.rack_input.currentText()),
                    "slot": int(self.slot_input.currentText()),
                    "simulate": self.sim_check.isChecked()
                },
                "profiles": self.profiles
            }
            with open(self.config_path, 'w') as f:
                json.dump(config, f, indent=4)
        except Exception as e:
            print(f"Error saving config: {e}")

    def update_rubrica_combo(self):
        self.rubrica_combo.blockSignals(True)
        self.rubrica_combo.clear()
        self.rubrica_combo.addItem("Seleziona Profilo...")
        for profile in self.profiles:
            self.rubrica_combo.addItem(profile["name"])
        self.rubrica_combo.blockSignals(False)

    def on_profile_selected(self, index):
        if index <= 0 or index > len(self.profiles):
            return
        profile = self.profiles[index - 1]
        
        # Load profile connection configurations
        if self.plc_client.is_connected():
            self.toggle_connection()
            
        self.ip_input.setText(profile["ip"])
        self.rack_input.setCurrentText(str(profile["rack"]))
        self.slot_input.setCurrentText(str(profile["slot"]))
        self.sim_check.setChecked(profile["simulate"])
        
        # Reset selection index
        self.rubrica_combo.blockSignals(True)
        self.rubrica_combo.setCurrentIndex(0)
        self.rubrica_combo.blockSignals(False)

    def save_current_profile(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "Salva Profilo", "Inserisci il nome per questo profilo PLC:")
        if not ok or not name.strip():
            return
            
        name = name.strip()
        existing = next((p for p in self.profiles if p["name"].lower() == name.lower()), None)
        profile = {
            "name": name,
            "ip": self.ip_input.text().strip(),
            "rack": int(self.rack_input.currentText()),
            "slot": int(self.slot_input.currentText()),
            "simulate": self.sim_check.isChecked()
        }
        
        if existing:
            reply = QMessageBox.question(
                self, "Sovrascrivi Profilo",
                f"Il profilo '{name}' esiste già. Vuoi sovrascriverlo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return
            self.profiles.remove(existing)
            
        self.profiles.append(profile)
        self.update_rubrica_combo()
        self.save_config()
        QMessageBox.information(self, "Profilo Salvato", f"Profilo '{name}' salvato correttamente in rubrica.")

    def manage_profiles(self):
        from PyQt6.QtWidgets import QDialog, QListWidget
        dialog = QDialog(self)
        dialog.setWindowTitle("Gestione Rubrica PLC")
        dialog.setMinimumSize(300, 250)
        
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("<b>Profili PLC salvati:</b>"))
        
        list_widget = QListWidget()
        for p in self.profiles:
            list_widget.addItem(p["name"])
        layout.addWidget(list_widget)
        
        btn_layout = QHBoxLayout()
        delete_btn = QPushButton("Elimina Selezionato")
        delete_btn.setStyleSheet("color: red;")
        btn_layout.addWidget(delete_btn)
        
        close_btn = QPushButton("Chiudi")
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        
        def delete_selected():
            selected_items = list_widget.selectedItems()
            if not selected_items:
                return
            name = selected_items[0].text()
            reply = QMessageBox.question(
                self, "Elimina Profilo",
                f"Sei sicuro di voler eliminare il profilo '{name}'?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                p_to_remove = next((p for p in self.profiles if p["name"] == name), None)
                if p_to_remove:
                    self.profiles.remove(p_to_remove)
                    self.update_rubrica_combo()
                    self.save_config()
                    list_widget.takeItem(list_widget.row(selected_items[0]))
                    
        delete_btn.clicked.connect(delete_selected)
        dialog.exec()

    def add_db_manually(self):
        from PyQt6.QtWidgets import QInputDialog
        db_num, ok = QInputDialog.getInt(self, "Aggiungi DB Manualmente", "Inserisci il numero del Data Block (DB):", 1, 1, 65535, 1)
        if not ok:
            return
            
        if db_num in self.dbs_list:
            QMessageBox.warning(self, "DB esistente", f"Il DB {db_num} è già presente nell'elenco.")
            return
            
        # Determine size
        size = 0
        if self.plc_client.is_connected():
            self.update_status_bar(f"Rilevamento dimensione per DB {db_num}...")
            try:
                size = self.plc_client.get_db_size(db_num)
            except Exception as e:
                # Fallback if size detection fails
                QMessageBox.information(
                    self, "Rilevamento Fallito",
                    f"Impossibile rilevare la dimensione del DB {db_num} dal PLC.\n"
                    "Inserisci la dimensione manualmente."
                )
                
        if size <= 0:
            # Ask user for size manually
            size, ok = QInputDialog.getInt(
                self, "Dimensione DB",
                f"Inserisci la dimensione del DB {db_num} in byte:",
                100, 1, 65535, 1
            )
            if not ok:
                return
                
        # Add to list
        self.dbs_list.append(db_num)
        self.dbs_list.sort()
        self.dbs_sizes[db_num] = size
        
        # Update UI
        self.populate_blocks_table()
        self.update_project_tree_online()
        self.update_status_bar(f"Aggiunto manualmente DB {db_num} con dimensione {size} byte.")

    def toggle_db_range_scan(self):
        # If already running, cancel it
        if hasattr(self, 'db_scan_worker') and self.db_scan_worker.isRunning():
            self.db_scan_worker.is_cancelled = True
            self.scan_range_btn.setText("Interruzione...")
            self.scan_range_btn.setEnabled(False)
            return

        # Check if connected
        if not self.plc_client.is_connected():
            QMessageBox.warning(self, "PLC non connesso", "Devi connetterti al PLC prima di poter scansionare l'intervallo DB.")
            return

        # Show ScanRangeDialog
        dialog = ScanRangeDialog(self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
            
        start_val = dialog.start_input.value()
        end_val = dialog.end_input.value()

        # Setup worker
        self.scan_range_btn.setText("Ferma Scansione")
        self.update_status_bar(f"Scansione DB avviata (da DB {start_val} a DB {end_val})...")
        
        self.db_scan_worker = DBScanWorker(self.plc_client, start_val, end_val)
        self.db_scan_worker.progress.connect(self.on_db_scan_progress)
        self.db_scan_worker.db_found.connect(self.on_db_found)
        self.db_scan_worker.finished.connect(self.on_db_scan_finished)
        self.db_scan_worker.start()

    def on_db_found(self, db_num, size):
        if db_num not in self.dbs_list:
            self.dbs_list.append(db_num)
            self.dbs_list.sort()
            self.dbs_sizes[db_num] = size
            self.populate_blocks_table()
            self.update_project_tree_online()
            self.update_status_bar(f"Trovato DB {db_num} ({size} byte)!")

    def on_db_scan_progress(self, completed, total):
        pct = int(completed / total * 100) if total > 0 else 0
        self.update_status_bar(f"Scansione DB in corso: {completed} di {total} verificati ({pct}%)...")

    def on_db_scan_finished(self, found_dbs):
        self.scan_range_btn.setText("Scansiona Intervallo DB...")
        self.scan_range_btn.setEnabled(True)
        self.update_status_bar(f"Scansione DB completata. Trovati {len(found_dbs)} DB nell'intervallo.")
        QMessageBox.information(self, "Scansione Completata", f"Scansione completata. Trovati {len(found_dbs)} blocchi DB.")
