from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QTableWidget, QTableWidgetItem, 
                             QComboBox, QLineEdit, QFileDialog, QMessageBox, 
                             QHeaderView, QCheckBox, QInputDialog)
from PyQt6.QtCore import QTimer, Qt
from PyQt6.QtGui import QIcon, QColor
import json

from utils import parse_s7_data, pack_s7_data, parse_s7_address
from ui.icons import get_custom_icon

class WatchTableViewer(QWidget):
    def __init__(self, parent=None, plc_client=None, watch_tables=None, on_tables_changed=None):
        super().__init__(parent)
        self.plc_client = plc_client
        self.watch_tables = watch_tables if watch_tables is not None else {}
        self.active_table_name = None
        self.on_tables_changed = on_tables_changed
        
        # Monitor timer
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.poll_plc)
        
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 1. Header Controls
        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel("<b>Tabella di Variabili:</b>"))
        
        self.table_selector = QComboBox()
        self.table_selector.setMinimumWidth(180)
        self.table_selector.currentIndexChanged.connect(self.on_table_selected)
        header_layout.addWidget(self.table_selector)
        
        self.new_table_btn = QPushButton("Nuova...")
        self.new_table_btn.setIcon(get_custom_icon("add"))
        self.new_table_btn.clicked.connect(self.create_new_table)
        header_layout.addWidget(self.new_table_btn)
        
        self.rename_table_btn = QPushButton("Rinomina...")
        self.rename_table_btn.clicked.connect(self.rename_current_table)
        header_layout.addWidget(self.rename_table_btn)
        
        self.delete_table_btn = QPushButton("Elimina")
        self.delete_table_btn.setIcon(get_custom_icon("delete"))
        self.delete_table_btn.clicked.connect(self.delete_current_table)
        header_layout.addWidget(self.delete_table_btn)
        
        header_layout.addStretch()
        
        # Live Monitor Controls
        self.monitor_checkbox = QCheckBox("Monitor Live (Attivo)")
        self.monitor_checkbox.stateChanged.connect(self.toggle_monitoring)
        self.monitor_checkbox.setEnabled(False)
        header_layout.addWidget(self.monitor_checkbox)
        
        header_layout.addWidget(QLabel("Intervallo (ms):"))
        self.interval_combo = QComboBox()
        self.interval_combo.addItems(["200", "500", "1000", "2000"])
        self.interval_combo.setCurrentIndex(1) # Default 500ms
        self.interval_combo.currentIndexChanged.connect(self.update_monitor_interval)
        header_layout.addWidget(self.interval_combo)
        
        layout.addLayout(header_layout)
        
        # 2. Secondary Toolbar Actions
        actions_layout = QHBoxLayout()
        self.import_btn = QPushButton("Importa Tabella (Excel)...")
        self.import_btn.setIcon(get_custom_icon("restore"))
        self.import_btn.clicked.connect(self.import_table)
        actions_layout.addWidget(self.import_btn)
        
        self.export_btn = QPushButton("Esporta Tabella (Excel)...")
        self.export_btn.setIcon(get_custom_icon("backup"))
        self.export_btn.clicked.connect(self.export_table)
        actions_layout.addWidget(self.export_btn)
        
        actions_layout.addStretch()
        
        self.add_var_btn = QPushButton("Aggiungi Variabile")
        self.add_var_btn.setIcon(get_custom_icon("add"))
        self.add_var_btn.clicked.connect(lambda: self.add_variable_row())
        actions_layout.addWidget(self.add_var_btn)
        
        layout.addLayout(actions_layout)
        
        # 3. Main Table Widget
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Nome / Simbolo", "Indirizzo (I, Q, M, DB)", "Tipo Dato", "Valore Live", "Nuovo Valore", "Azione"
        ])
        
        # Interactive Resizing on Header
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 220)
        self.table.setColumnWidth(1, 140)
        self.table.setColumnWidth(2, 100)
        self.table.setColumnWidth(3, 120)
        self.table.setColumnWidth(4, 120)
        self.table.setColumnWidth(5, 110)
        
        self.table.itemChanged.connect(self.on_table_item_changed)
        layout.addWidget(self.table)
        
        # 4. Bottom Read/Write Actions
        bottom_layout = QHBoxLayout()
        self.read_all_btn = QPushButton("Leggi Tutte le Variabili")
        self.read_all_btn.clicked.connect(self.read_all_variables)
        bottom_layout.addWidget(self.read_all_btn)
        
        self.write_all_btn = QPushButton("Scrivi Valori Inseriti")
        self.write_all_btn.clicked.connect(self.write_all_variables)
        bottom_layout.addWidget(self.write_all_btn)
        
        layout.addLayout(bottom_layout)
        
        self.refresh_table_list()
        
    def set_watch_tables(self, watch_tables):
        self.watch_tables = watch_tables if watch_tables is not None else {}
        self.refresh_table_list()
        
    def set_plc_client(self, plc_client):
        self.plc_client = plc_client
        self.monitor_checkbox.setEnabled(self.plc_client is not None and self.plc_client.is_connected())
        
    def refresh_table_list(self, select_table=None):
        self.table_selector.blockSignals(True)
        self.table_selector.clear()
        
        if not self.watch_tables:
            # Ensure at least one default table exists
            self.watch_tables["Tabella_1"] = [
                {"name": "Start_Pulsante", "address": "I0.0", "type": "BOOL"},
                {"name": "Motore_Uscita", "address": "Q0.0", "type": "BOOL"},
                {"name": "Velocita_Set", "address": "MW10", "type": "INT"},
                {"name": "Temperatura", "address": "DB1.DBD0", "type": "REAL"}
            ]
            if self.on_tables_changed:
                self.on_tables_changed(self.watch_tables)
                
        for tname in self.watch_tables.keys():
            self.table_selector.addItem(tname)
            
        target = select_table if select_table and select_table in self.watch_tables else list(self.watch_tables.keys())[0]
        self.table_selector.setCurrentText(target)
        self.table_selector.blockSignals(False)
        
        self.load_active_table(target)
        
    def on_table_selected(self, index):
        tname = self.table_selector.currentText()
        if tname and tname in self.watch_tables:
            self.load_active_table(tname)
            
    def load_active_table(self, table_name):
        self.active_table_name = table_name
        self.table.blockSignals(True)
        self.table.setRowCount(0)
        
        vars_list = self.watch_tables.get(table_name, [])
        for var in vars_list:
            self._insert_row_ui(var.get("name", "Var"), var.get("address", "I0.0"), var.get("type", "BOOL"))
            
        self.table.blockSignals(False)
        self.update_live_values()
        
    def create_new_table(self, default_name=None):
        name = default_name
        if not name:
            name, ok = QInputDialog.getText(self, "Nuova Tabella di Variabili", "Nome tabella:")
            if not ok or not name.strip():
                return
            name = name.strip()
            
        if name in self.watch_tables:
            QMessageBox.warning(self, "Nome Esistente", f"La tabella '{name}' esiste già.")
            return
            
        self.watch_tables[name] = [
            {"name": "Var_1", "address": "I0.0", "type": "BOOL"}
        ]
        if self.on_tables_changed:
            self.on_tables_changed(self.watch_tables)
            
        self.refresh_table_list(select_table=name)
        
    def rename_current_table(self, old_name=None, new_name=None):
        target_old = old_name or self.active_table_name
        if not target_old or target_old not in self.watch_tables:
            return
            
        if not new_name:
            new_name, ok = QInputDialog.getText(self, "Rinomina Tabella", "Nuovo nome:", text=target_old)
            if not ok or not new_name.strip() or new_name.strip() == target_old:
                return
            new_name = new_name.strip()
            
        if new_name in self.watch_tables:
            QMessageBox.warning(self, "Nome Esistente", f"Esiste già una tabella denominata '{new_name}'.")
            return
            
        self.watch_tables[new_name] = self.watch_tables.pop(target_old)
        if self.on_tables_changed:
            self.on_tables_changed(self.watch_tables)
            
        self.refresh_table_list(select_table=new_name)
        
    def delete_current_table(self, target_name=None):
        tname = target_name or self.active_table_name
        if not tname or tname not in self.watch_tables:
            return
            
        if len(self.watch_tables) <= 1:
            QMessageBox.warning(self, "Eliminazione Impossibile", "Devi mantenere almeno una tabella di variabili.")
            return
            
        reply = QMessageBox.question(
            self, "Conferma Eliminazione",
            f"Sei sicuro di voler eliminare la tabella '{tname}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.No:
            return
            
        del self.watch_tables[tname]
        if self.on_tables_changed:
            self.on_tables_changed(self.watch_tables)
            
        self.refresh_table_list()
        
    def add_variable_row(self, name="Nuova_Var", address="I0.0", dtype="BOOL"):
        if not self.active_table_name:
            return
            
        self.table.blockSignals(True)
        self._insert_row_ui(name, address, dtype)
        
        # Sync model list
        vars_list = self.watch_tables.setdefault(self.active_table_name, [])
        vars_list.append({"name": name, "address": address, "type": dtype})
        
        self.table.blockSignals(False)
        
        if self.on_tables_changed:
            self.on_tables_changed(self.watch_tables)
            
        self.update_live_values()

    def _insert_row_ui(self, name, address, dtype):
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Col 0: Name / Symbol
        name_item = QTableWidgetItem(name)
        self.table.setItem(row, 0, name_item)
        
        # Col 1: Address string
        addr_item = QTableWidgetItem(address)
        parsed = parse_s7_address(address)
        if not parsed["valid"]:
            addr_item.setForeground(QColor("red"))
            addr_item.setToolTip(parsed.get("error", "Indirizzo non valido"))
        self.table.setItem(row, 1, addr_item)
        
        # Col 2: Datatype ComboBox
        type_combo = QComboBox()
        type_combo.addItems(["BOOL", "BYTE", "CHAR", "INT", "WORD", "DINT", "DWORD", "REAL", "STRING"])
        type_combo.setCurrentText(dtype)
        type_combo.setProperty("row", row)
        type_combo.currentIndexChanged.connect(self.on_type_combo_changed)
        self.table.setCellWidget(row, 2, type_combo)
        
        # Col 3: Live Value (Read-Only)
        live_item = QTableWidgetItem("---")
        live_item.setFlags(live_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row, 3, live_item)
        
        # Col 4: New Value Input LineEdit
        write_input = QLineEdit()
        write_input.setProperty("row", row)
        self.table.setCellWidget(row, 4, write_input)
        
        # Col 5: Action Widget (Write & Delete Buttons)
        act_widget = QWidget()
        act_layout = QHBoxLayout(act_widget)
        act_layout.setContentsMargins(2, 2, 2, 2)
        act_layout.setSpacing(4)
        
        write_btn = QPushButton()
        write_btn.setIcon(get_custom_icon("write"))
        write_btn.setToolTip("Scrivi valore sul PLC")
        write_btn.clicked.connect(lambda: self.write_single_variable(row))
        act_layout.addWidget(write_btn)
        
        del_btn = QPushButton()
        del_btn.setIcon(get_custom_icon("delete"))
        del_btn.setToolTip("Elimina variabile")
        del_btn.clicked.connect(lambda: self.delete_variable_row(row))
        act_layout.addWidget(del_btn)
        
        self.table.setCellWidget(row, 5, act_widget)

    def delete_variable_row(self, row):
        if not self.active_table_name or row < 0 or row >= self.table.rowCount():
            return
            
        self.table.blockSignals(True)
        self.table.removeRow(row)
        
        vars_list = self.watch_tables.get(self.active_table_name, [])
        if 0 <= row < len(vars_list):
            vars_list.pop(row)
            
        # Re-index properties
        for r in range(self.table.rowCount()):
            combo = self.table.cellWidget(r, 2)
            if combo: combo.setProperty("row", r)
            line = self.table.cellWidget(r, 4)
            if line: line.setProperty("row", r)
            
        self.table.blockSignals(False)
        
        if self.on_tables_changed:
            self.on_tables_changed(self.watch_tables)

    def on_table_item_changed(self, item):
        row = item.row()
        col = item.column()
        if not self.active_table_name or row < 0 or col not in [0, 1]:
            return
            
        vars_list = self.watch_tables.get(self.active_table_name, [])
        if row >= len(vars_list):
            return
            
        val = item.text().strip()
        
        if col == 0:  # Name
            vars_list[row]["name"] = val
        elif col == 1:  # Address
            vars_list[row]["address"] = val
            parsed = parse_s7_address(val)
            if parsed["valid"]:
                item.setForeground(QColor("black"))
                item.setToolTip("")
                # Auto predict datatype if valid
                auto_type = parsed["datatype"]
                combo = self.table.cellWidget(row, 2)
                if combo:
                    combo.blockSignals(True)
                    combo.setCurrentText(auto_type)
                    combo.blockSignals(False)
                vars_list[row]["type"] = auto_type
            else:
                item.setForeground(QColor("red"))
                item.setToolTip(parsed.get("error", "Sintassi errata"))
                
        if self.on_tables_changed:
            self.on_tables_changed(self.watch_tables)
            
        self.update_live_values()

    def on_type_combo_changed(self, idx):
        sender = self.sender()
        if not sender: return
        row = sender.property("row")
        if not self.active_table_name or row is None: return
        
        vars_list = self.watch_tables.get(self.active_table_name, [])
        if 0 <= row < len(vars_list):
            vars_list[row]["type"] = sender.currentText()
            if self.on_tables_changed:
                self.on_tables_changed(self.watch_tables)
            self.update_live_values()

    def toggle_monitoring(self, state):
        if state == 2:  # Checked
            interval = int(self.interval_combo.currentText())
            self.timer.start(interval)
            self.read_all_btn.setEnabled(False)
            self.write_all_btn.setEnabled(False)
        else:
            self.timer.stop()
            self.read_all_btn.setEnabled(True)
            self.write_all_btn.setEnabled(True)

    def update_monitor_interval(self, idx):
        if self.timer.isActive():
            interval = int(self.interval_combo.currentText())
            self.timer.start(interval)

    def poll_plc(self):
        if not self.plc_client or not self.plc_client.is_connected():
            self.monitor_checkbox.setChecked(False)
            return
        self.update_live_values()

    def read_all_variables(self):
        if not self.plc_client or not self.plc_client.is_connected():
            QMessageBox.warning(self, "Non Connesso", "PLC non connesso.")
            return
        self.update_live_values()

    def update_live_values(self):
        is_conn = self.plc_client is not None and self.plc_client.is_connected()
        vars_list = self.watch_tables.get(self.active_table_name, [])
        
        self.table.blockSignals(True)
        try:
            for row, var in enumerate(vars_list):
                if row >= self.table.rowCount():
                    continue
                    
                addr_str = var.get("address", "")
                dtype = var.get("type", "INT")
                parsed = parse_s7_address(addr_str)
                
                val = "---"
                if is_conn and parsed["valid"]:
                    area = parsed["area_code"]
                    db_num = parsed["db_number"]
                    byte_off = parsed["byte_offset"]
                    bit_off = parsed["bit_offset"]
                    
                    # Determine required read length
                    size = 1
                    if dtype in ["INT", "WORD"]: size = 2
                    elif dtype in ["DINT", "DWORD", "REAL"]: size = 4
                    elif dtype == "STRING": size = 256
                    
                    try:
                        data_bytes = self.plc_client.read_area_bytes(area, db_num, byte_off, size)
                        val = parse_s7_data(dtype, data_bytes, 0, bit_off)
                        if val is None:
                            val = "---"
                    except Exception:
                        val = "Err"
                        
                item = self.table.item(row, 3)
                if item:
                    item.setText(str(val))
                    if val in ["---", "Err"]:
                        item.setForeground(QColor("gray") if val == "---" else QColor("red"))
                    else:
                        item.setForeground(QColor("black"))
        finally:
            self.table.blockSignals(False)

    def write_single_variable(self, row):
        if not self.plc_client or not self.plc_client.is_connected():
            QMessageBox.warning(self, "Non Connesso", "PLC non connesso.")
            return
            
        vars_list = self.watch_tables.get(self.active_table_name, [])
        if row < 0 or row >= len(vars_list):
            return
            
        var = vars_list[row]
        line_edit = self.table.cellWidget(row, 4)
        if not line_edit:
            return
            
        val_str = line_edit.text().strip()
        if not val_str:
            QMessageBox.warning(self, "Valore Vuoto", "Inserisci un valore da scrivere.")
            return
            
        parsed = parse_s7_address(var["address"])
        if not parsed["valid"]:
            QMessageBox.critical(self, "Indirizzo Errato", f"Indirizzo '{var['address']}' non valido.")
            return
            
        dtype = var["type"]
        area = parsed["area_code"]
        db_num = parsed["db_number"]
        byte_off = parsed["byte_offset"]
        bit_off = parsed["bit_offset"]
        
        # Parse user value
        try:
            if dtype == "BOOL":
                val_to_write = True if val_str.lower() in ["true", "1", "on", "yes"] else False
            elif dtype in ["BYTE", "INT", "WORD", "DINT", "DWORD"]:
                val_to_write = int(val_str)
            elif dtype == "REAL":
                val_to_write = float(val_str)
            else:
                val_to_write = val_str
        except ValueError:
            QMessageBox.critical(self, "Valore Errato", f"Impossibile convertire '{val_str}' nel tipo {dtype}.")
            return
            
        try:
            if dtype == "BOOL":
                # Read byte to preserve other bits
                existing_bytes = self.plc_client.read_area_bytes(area, db_num, byte_off, 1)
                packed = pack_s7_data("BOOL", val_to_write, 0, bit_off, existing_bytes)
            else:
                packed = pack_s7_data(dtype, val_to_write, 0, 0)
                
            if packed is None:
                raise ValueError("Errore nella codifica del valore.")
                
            self.plc_client.write_area_bytes(area, db_num, byte_off, packed)
            line_edit.clear()
            self.update_live_values()
            
        except Exception as e:
            QMessageBox.critical(self, "Errore Scrittura", f"Impossibile scrivere la variabile:\n{str(e)}")

    def write_all_variables(self):
        if not self.plc_client or not self.plc_client.is_connected():
            QMessageBox.warning(self, "Non Connesso", "PLC non connesso.")
            return
            
        count = 0
        for row in range(self.table.rowCount()):
            line = self.table.cellWidget(row, 4)
            if line and line.text().strip():
                count += 1
                
        if count == 0:
            QMessageBox.information(self, "Nessun Valore Inserito", "Inserisci i valori da scrivere nella colonna 'Nuovo Valore'.")
            return
            
        reply = QMessageBox.question(
            self, "Conferma Scrittura",
            f"Confermi la scrittura di {count} variabili sul PLC?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.No:
            return
            
        for row in range(self.table.rowCount()):
            line = self.table.cellWidget(row, 4)
            if line and line.text().strip():
                self.write_single_variable(row)

    def export_table(self):
        import os
        if not self.active_table_name:
            return
        vars_list = self.watch_tables.get(self.active_table_name, [])
        filepath, selected_filter = QFileDialog.getSaveFileName(
            self, "Esporta Tabella Variabili", f"{self.active_table_name}.xlsx", "File Excel (*.xlsx);;File JSON (*.json)"
        )
        if not filepath:
            return
            
        try:
            if filepath.endswith('.json'):
                with open(filepath, 'w') as f:
                    json.dump(vars_list, f, indent=4)
            else:
                if not filepath.endswith('.xlsx'):
                    filepath += '.xlsx'
                    
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = self.active_table_name[:30].replace("/", "_").replace("\\", "_")
                
                # Styles
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
                regular_font = Font(name="Segoe UI", size=10)
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                headers = ["Nome / Simbolo", "Indirizzo (I, Q, M, DB)", "Tipo Dato", "Valore Live"]
                ws.append(headers)
                ws.row_dimensions[1].height = 25
                
                for col_idx in range(1, 5):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border
                    
                for row_idx, var in enumerate(vars_list, start=2):
                    live_item = self.table.item(row_idx - 2, 3)
                    live_val = live_item.text() if live_item else "---"
                    row_data = [var.get("name", ""), var.get("address", ""), var.get("type", ""), live_val]
                    ws.append(row_data)
                    ws.row_dimensions[row_idx].height = 20
                    
                    for col_idx in range(1, 5):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = regular_font
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                # Auto-fit columns
                for col in ws.columns:
                    max_len = 0
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    for cell in col:
                        val_str = str(cell.value or "")
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
                    
                wb.save(filepath)
                
            QMessageBox.information(self, "Esportazione Completata", f"Tabella '{self.active_table_name}' esportata con successo.")
        except Exception as e:
            QMessageBox.critical(self, "Errore Esportazione", f"Impossibile esportare la tabella:\n{str(e)}")

    def import_table(self):
        import os
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Importa Tabella Variabili", "", "File Excel / JSON (*.xlsx *.json);;File Excel (*.xlsx);;File JSON (*.json)"
        )
        if not filepath:
            return
            
        try:
            imported_vars = []
            if filepath.endswith('.json'):
                with open(filepath, 'r') as f:
                    data = json.load(f)
                if not isinstance(data, list):
                    raise ValueError("Il file JSON deve contenere una lista di variabili.")
                imported_vars = data
            else:
                import openpyxl
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
                
                # Detect column mapping from header row
                col_name = 1
                col_addr = 2
                col_type = 3
                start_row = 1
                
                # Check if row 1 contains header labels
                row1_vals = [str(ws.cell(row=1, column=c).value or "").strip().upper() for c in range(1, max(ws.max_column + 1, 4))]
                has_header = False
                for c_idx, val in enumerate(row1_vals, start=1):
                    if any(k in val for k in ["NOME", "NAME", "SIMBOLO", "SYMBOL", "VAR"]):
                        col_name = c_idx
                        has_header = True
                    elif any(k in val for k in ["INDIRIZZO", "ADDRESS", "OFFSET"]):
                        col_addr = c_idx
                        has_header = True
                    elif any(k in val for k in ["TIPO", "TYPE", "DATA"]):
                        col_type = c_idx
                        has_header = True
                        
                if has_header:
                    start_row = 2
                    
                for r_idx in range(start_row, ws.max_row + 1):
                    val_name = str(ws.cell(row=r_idx, column=col_name).value or "").strip()
                    val_addr = str(ws.cell(row=r_idx, column=col_addr).value or "").strip()
                    val_type = str(ws.cell(row=r_idx, column=col_type).value or "").strip()
                    
                    if not val_addr:
                        # Skip empty address rows
                        continue
                        
                    parsed = parse_s7_address(val_addr)
                    if not val_name:
                        val_name = f"Var_{r_idx}"
                    if not val_type or val_type not in ["BOOL", "BYTE", "CHAR", "INT", "WORD", "DINT", "DWORD", "REAL", "STRING"]:
                        val_type = parsed["datatype"] if parsed["valid"] else "INT"
                        
                    imported_vars.append({
                        "name": val_name,
                        "address": val_addr,
                        "type": val_type
                    })
                    
            if not imported_vars:
                raise ValueError("Nessuna variabile valida trovata nel file.")
                
            base_name = os.path.splitext(os.path.basename(filepath))[0]
            table_name = base_name
            count = 1
            while table_name in self.watch_tables:
                table_name = f"{base_name}_{count}"
                count += 1
                
            self.watch_tables[table_name] = imported_vars
            if self.on_tables_changed:
                self.on_tables_changed(self.watch_tables)
                
            self.refresh_table_list(select_table=table_name)
            QMessageBox.information(self, "Importazione Completata", f"Importate {len(imported_vars)} variabili nella tabella '{table_name}'.")
        except Exception as e:
            QMessageBox.critical(self, "Errore Importazione", f"Impossibile importare la tabella:\n{str(e)}")
