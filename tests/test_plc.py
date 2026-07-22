import sys
import os
import unittest

# Add src folder to python path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../src')))

from plc_comm import PLCClient
from utils import parse_s7_data, pack_s7_data, parse_s7_address

class TestS7BackupTool(unittest.TestCase):
    def setUp(self):
        # Initialize client in simulation mode for local offline unit testing
        self.client = PLCClient(simulate=True)
        self.client.connect("127.0.0.1")

    def tearDown(self):
        self.client.disconnect()

    def test_connection_and_cpu_info(self):
        self.assertTrue(self.client.is_connected())
        info = self.client.get_cpu_info()
        self.assertEqual(info["ModuleTypeName"], "CPU 315-2 PN/DP (SIM)")
        self.assertEqual(info["SerialNumber"], "S7-SIM-12345678")

    def test_list_dbs(self):
        dbs = self.client.list_dbs()
        self.assertIn(1, dbs)
        self.assertIn(2, dbs)
        self.assertIn(10, dbs)
        self.assertIn(100, dbs)

    def test_db_read_write(self):
        # Read initial simulated DB10 data (should be all zeros)
        size = self.client.get_db_size(10)
        self.assertEqual(size, 50)
        
        data = self.client.read_db_bytes(10, size)
        self.assertEqual(len(data), 50)
        self.assertEqual(data[0], 0)
        
        # Modify some bytes
        data[0] = 0xAA
        data[1] = 0x55
        self.client.write_db_bytes(10, data)
        
        # Read back and check
        data_read = self.client.read_db_bytes(10, size)
        self.assertEqual(data_read[0], 0xAA)
        self.assertEqual(data_read[1], 0x55)

    def test_s7_datatype_parsing(self):
        buffer = bytearray(20)
        
        # Test INT packing and parsing (Big Endian)
        pack_s7_data("INT", 1234, 0, existing_bytes=buffer)
        val_int = parse_s7_data("INT", buffer, 0)
        self.assertEqual(val_int, 1234)
        
        # Test REAL packing and parsing (Big Endian float)
        pack_s7_data("REAL", 3.1415, 2, existing_bytes=buffer)
        val_real = parse_s7_data("REAL", buffer, 2)
        self.assertAlmostEqual(val_real, 3.1415, places=4)
        
        # Test BOOL packing and parsing
        pack_s7_data("BOOL", True, 6, bit_offset=2, existing_bytes=buffer)
        pack_s7_data("BOOL", False, 6, bit_offset=3, existing_bytes=buffer)
        
        val_bool_2 = parse_s7_data("BOOL", buffer, 6, bit_offset=2)
        val_bool_3 = parse_s7_data("BOOL", buffer, 6, bit_offset=3)
        self.assertTrue(val_bool_2)
        self.assertFalse(val_bool_3)
        
        # Test STRING packing and parsing (Siemens format: [max_len, curr_len, chars...])
        buffer_str = bytearray([10, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
        pack_s7_data("STRING", "HELLO", 0, existing_bytes=buffer_str)
        val_str = parse_s7_data("STRING", buffer_str, 0)
        self.assertEqual(val_str, "HELLO")

    def test_address_parsing(self):
        # Inputs (I)
        res_i = parse_s7_address("I0.4")
        self.assertTrue(res_i["valid"])
        self.assertEqual(res_i["area_code"], 0x81)
        self.assertEqual(res_i["byte_offset"], 0)
        self.assertEqual(res_i["bit_offset"], 4)
        self.assertEqual(res_i["datatype"], "BOOL")

        # Outputs (Q)
        res_q = parse_s7_address("Q10.2")
        self.assertTrue(res_q["valid"])
        self.assertEqual(res_q["area_code"], 0x82)
        self.assertEqual(res_q["byte_offset"], 10)
        self.assertEqual(res_q["bit_offset"], 2)

        # Merkers (M)
        res_m = parse_s7_address("MW10")
        self.assertTrue(res_m["valid"])
        self.assertEqual(res_m["area_code"], 0x83)
        self.assertEqual(res_m["byte_offset"], 10)
        self.assertEqual(res_m["datatype"], "INT")

        # DB
        res_db = parse_s7_address("DB1.DBD4")
        self.assertTrue(res_db["valid"])
        self.assertEqual(res_db["area_code"], 0x84)
        self.assertEqual(res_db["db_number"], 1)
        self.assertEqual(res_db["byte_offset"], 4)
        self.assertEqual(res_db["datatype"], "REAL")

    def test_area_read_write(self):
        # Write to simulated merker area (MK 0x83)
        packed_bytes = pack_s7_data("INT", 5678, 0)
        self.client.write_area_bytes(0x83, 0, 10, packed_bytes)
        
        # Read back
        read_bytes = self.client.read_area_bytes(0x83, 0, 10, 2)
        val = parse_s7_data("INT", read_bytes, 0)
        self.assertEqual(val, 5678)

    def test_excel_watch_table(self):
        import openpyxl
        import tempfile
        
        # Create temp excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
            temp_path = tf.name
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Nome / Simbolo", "Indirizzo (I, Q, M, DB)", "Tipo Dato"])
            ws.append(["Pulsante_Start", "I0.0", "BOOL"])
            ws.append(["Motore_Run", "Q0.1", "BOOL"])
            ws.append(["Pressione", "DB1.DBD0", "REAL"])
            wb.save(temp_path)
            
            # Read back workbook
            wb_read = openpyxl.load_workbook(temp_path, data_only=True)
            ws_read = wb_read.active
            self.assertEqual(ws_read.cell(row=2, column=1).value, "Pulsante_Start")
            self.assertEqual(ws_read.cell(row=2, column=2).value, "I0.0")
            self.assertEqual(ws_read.cell(row=4, column=2).value, "DB1.DBD0")
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

if __name__ == '__main__':
    unittest.main()
